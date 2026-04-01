#!/usr/bin/env python3
"""
One-time setup script to extract column headers from the Zielformat Excel template.
Run this once before starting the n8n workflow for the first time.

Usage:
    python3 extract_headers.py

Output:
    files/zielformat-headers.json
"""

import json
import os
from openpyxl import load_workbook

ZIELFORMAT_PATH = "Ziel-Tabelle.xlsx"
OUTPUT_PATH = "files/zielformat-headers.json"

def extract_headers():
    if not os.path.exists(ZIELFORMAT_PATH):
        print(f"Error: {ZIELFORMAT_PATH} not found in current directory")
        return

    os.makedirs("files", exist_ok=True)

    wb = load_workbook(ZIELFORMAT_PATH, read_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(max_row=1)) if cell.value]

    with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
        json.dump(headers, f, ensure_ascii=False, indent=2)

    print(f"Successfully extracted {len(headers)} headers to {OUTPUT_PATH}")

if __name__ == "__main__":
    extract_headers()